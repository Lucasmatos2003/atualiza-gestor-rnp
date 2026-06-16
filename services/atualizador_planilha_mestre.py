import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import process, fuzz
from datetime import datetime


LIMITE_NOME_ENCONTRADO = 90
LIMITE_NOME_REVISAR = 80
LIMITE_INEP_COM_NOME = 70


def normalizar_texto(texto):
    if texto is None or pd.isna(texto):
        return ""

    texto = str(texto).strip().upper()

    texto = unicodedata.normalize("NFD", texto)
    texto = texto.encode("ascii", "ignore")
    texto = texto.decode("utf-8")

    texto = texto.replace(".", " ")
    texto = texto.replace("-", " ")
    texto = texto.replace("_", " ")
    texto = texto.replace("/", " ")

    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def normalizar_chave_cidade(texto):
    texto = normalizar_texto(texto)
    texto = re.sub(r"[^A-Z0-9]", "", texto)

    return texto


def limpar_inep(valor):
    if valor is None or pd.isna(valor):
        return ""

    valor = re.sub(r"\D", "", str(valor))

    return valor


def limpar_telefone(valor):
    if valor is None or pd.isna(valor):
        return ""

    valor = str(valor)

    telefones = re.findall(
        r"(?:\+?55\s*)?(?:\(?\d{2}\)?\s*)?\d?\s?\d{4,5}[-\s]?\d{4}",
        valor
    )

    candidatos = []

    for telefone in telefones:
        numero = re.sub(r"\D", "", telefone)

        if numero.startswith("55") and len(numero) in [12, 13]:
            numero = numero[2:]

        if numero.startswith("0") and len(numero) in [11, 12]:
            numero = numero[1:]

        if len(numero) in [8, 9, 10, 11]:
            candidatos.append(numero)

    if candidatos:
        return candidatos[0]

    valor = re.sub(r"\D", "", valor)

    if valor.startswith("55") and len(valor) in [12, 13]:
        valor = valor[2:]

    if valor.startswith("0") and len(valor) in [11, 12]:
        valor = valor[1:]

    if len(valor) in [8, 9, 10, 11]:
        return valor

    return ""


def limpar_email(valor):
    if valor is None or pd.isna(valor):
        return ""

    valor = str(valor)

    valor = valor.replace("\n", " ")
    valor = re.sub(r"\s+", " ", valor)
    valor = re.sub(r"\s*@\s*", "@", valor)
    valor = re.sub(r"\s*\.\s*", ".", valor)

    emails = re.findall(
        r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
        valor
    )

    if emails:
        return emails[0].lower()

    return ""


def limpar_valor(valor):
    if valor is None or pd.isna(valor):
        return ""

    valor = str(valor).replace("\xa0", " ").strip()
    valor = re.sub(r"\s+", " ", valor)

    if valor.lower() in [
        "",
        "nan",
        "none",
        "-",
        "não informado",
        "nao informado",
        "não possui",
        "nao possui",
        "n/a"
    ]:
        return ""

    return valor


def coletar_telefones_contato(contato):
    telefones = []

    colunas = [
        "telefone",
        "telefone_2",
        "telefone_3",
        "outros_telefones"
    ]

    for coluna in colunas:
        valor = contato.get(coluna)

        if valor is None or pd.isna(valor):
            continue

        partes = re.split(r"[;/,|]", str(valor))

        for parte in partes:
            telefone = limpar_telefone(parte)

            if telefone:
                telefones.append(telefone)

    telefones_unicos = []

    for telefone in telefones:
        if telefone not in telefones_unicos:
            telefones_unicos.append(telefone)

    def prioridade(telefone):
        if len(telefone) == 11 and telefone[2] == "9":
            return 4

        if len(telefone) == 11:
            return 3

        if len(telefone) == 10:
            return 2

        if len(telefone) == 9:
            return 1

        return 0

    telefones_unicos = sorted(
        telefones_unicos,
        key=prioridade,
        reverse=True
    )

    return telefones_unicos


def limpar_nome_escola(nome):
    nome = normalizar_texto(nome)

    termos_remover = [
        r"\bESCOLA MUNICIPAL DE ENSINO FUNDAMENTAL\b",
        r"\bESCOLA MUNICIPAL DE EDUCACAO INFANTIL\b",
        r"\bESCOLA ESTADUAL DE ENSINO FUNDAMENTAL E MEDIO\b",
        r"\bESCOLA ESTADUAL DE ENSINO MEDIO\b",
        r"\bESCOLA DE REFERENCIA EM ENSINO MEDIO\b",
        r"\bESCOLA DE REFERENCIA EM ENSINO FUNDAMENTAL E MEDIO\b",
        r"\bESCOLA TECNICA ESTADUAL\b",
        r"\bESCOLA MUNICIPAL\b",
        r"\bESCOLA ESTADUAL\b",
        r"\bESCOLA\b",
        r"\bCENTRO MUNICIPAL DE EDUCACAO INFANTIL\b",
        r"\bCENTRO DE EDUCACAO INFANTIL\b",
        r"\bUNIDADE DE ACOLHIMENTO A CRIANCA DO PROGRAMA NOVA SEMENTE\b",
        r"\bUNIDADE DE ACOLHIMENTO A CRIANÇA DO PROGRAMA NOVA SEMENTE\b",
        r"\bEMEIEF\b",
        r"\bEMEIF\b",
        r"\bEMEITI\b",
        r"\bEMEF\b",
        r"\bEMEI\b",
        r"\bEEEFM\b",
        r"\bEEEF\b",
        r"\bEEEM\b",
        r"\bEEM\b",
        r"\bEE\b",
        r"\bEM\b",
        r"\bCMEI\b",
        r"\bCM\b",
        r"\bEREM\b",
        r"\bETE\b",
        r"\bETI\b",
        r"\bECI\b",
        r"\bECIT\b",
        r"\bCAIC\b"
    ]

    for termo in termos_remover:
        nome = re.sub(termo, " ", nome)

    nome = re.sub(r"\s+", " ", nome)

    return nome.strip()


def localizar_aba_mestre(wb):
    if "ESCOLAS_514" in wb.sheetnames:
        return wb["ESCOLAS_514"]

    return wb[wb.sheetnames[0]]


def localizar_cabecalho(ws):
    for row in range(1, min(ws.max_row, 20) + 1):
        valores = []

        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=row, column=col).value

            if valor:
                valores.append(normalizar_texto(valor))

        texto = " ".join(valores)

        if "INEP" in texto and "ESCOLAS" in texto and "GESTOR" in texto:
            return row

    return 1


def mapear_colunas_mestre(ws, linha_cabecalho):
    mapa = {}

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col).value

        if not valor:
            continue

        nome = normalizar_texto(valor)

        if nome == "INEP":
            mapa["inep"] = col

        elif nome in ["ESCOLAS", "ESCOLA", "UNIDADE ESCOLAR"]:
            mapa["escola"] = col

        elif nome == "REDE":
            mapa["rede"] = col

        elif nome == "CIDADE":
            mapa["cidade"] = col

        elif nome == "GESTOR":
            mapa["gestor"] = col

        elif nome in ["NUMERO DE TELEFONE", "NÚMERO DE TELEFONE"] or (
            "TELEFONE" in nome
            and "2" not in nome
            and "3" not in nome
            and "OUTROS" not in nome
        ):
            mapa["telefone"] = col

        elif nome in ["EMAIL", "E-MAIL", "E MAIL"]:
            mapa["email"] = col

        elif nome == "TELEFONE 2":
            mapa["telefone_2"] = col

        elif nome == "TELEFONE 3":
            mapa["telefone_3"] = col

        elif nome == "OUTROS TELEFONES":
            mapa["outros_telefones"] = col

        elif nome == "FONTE DO CONTATO":
            mapa["fonte_contato"] = col

        elif nome in ["DATA DA ATUALIZACAO", "DATA DA ATUALIZAÇÃO"]:
            mapa["data_atualizacao"] = col

        elif nome in ["STATUS DA ATUALIZACAO", "STATUS DA ATUALIZAÇÃO"]:
            mapa["status_atualizacao"] = col

    colunas_obrigatorias = [
        "inep",
        "escola",
        "cidade",
        "gestor",
        "telefone",
        "email"
    ]

    faltando = [
        coluna
        for coluna in colunas_obrigatorias
        if coluna not in mapa
    ]

    if faltando:
        raise ValueError(
            "A planilha mestre não possui as colunas obrigatórias: "
            + ", ".join(faltando)
        )

    return mapa


def garantir_coluna(ws, mapa, chave, nome_coluna, linha_cabecalho):
    if chave in mapa:
        return mapa

    nova_coluna = ws.max_column + 1

    ws.cell(
        row=linha_cabecalho,
        column=nova_coluna
    ).value = nome_coluna

    mapa[chave] = nova_coluna

    return mapa


def garantir_colunas_extras(ws, mapa, linha_cabecalho):
    mapa = garantir_coluna(
        ws,
        mapa,
        "telefone_2",
        "TELEFONE 2",
        linha_cabecalho
    )

    mapa = garantir_coluna(
        ws,
        mapa,
        "telefone_3",
        "TELEFONE 3",
        linha_cabecalho
    )

    mapa = garantir_coluna(
        ws,
        mapa,
        "outros_telefones",
        "OUTROS TELEFONES",
        linha_cabecalho
    )

    mapa = garantir_coluna(
        ws,
        mapa,
        "fonte_contato",
        "FONTE DO CONTATO",
        linha_cabecalho
    )

    mapa = garantir_coluna(
        ws,
        mapa,
        "data_atualizacao",
        "DATA DA ATUALIZAÇÃO",
        linha_cabecalho
    )

    mapa = garantir_coluna(
        ws,
        mapa,
        "status_atualizacao",
        "STATUS DA ATUALIZAÇÃO",
        linha_cabecalho
    )

    return mapa


def carregar_base_mestre(ws, mapa, linha_cabecalho):
    registros = []

    for row in range(linha_cabecalho + 1, ws.max_row + 1):
        inep = ws.cell(row=row, column=mapa["inep"]).value
        escola = ws.cell(row=row, column=mapa["escola"]).value
        cidade = ws.cell(row=row, column=mapa["cidade"]).value

        if not escola:
            continue

        registros.append({
            "linha_excel": row,
            "inep": limpar_inep(inep),
            "escola": escola,
            "escola_limpa": limpar_nome_escola(escola),
            "cidade": cidade,
            "cidade_limpa": normalizar_texto(cidade),
            "cidade_chave": normalizar_chave_cidade(cidade)
        })

    return pd.DataFrame(registros)


def buscar_por_inep_com_validacao(
    df_mestre_filtrado,
    inep_contato,
    escola_contato_limpa
):
    if not inep_contato:
        return None, None, None, None

    resultado_inep = df_mestre_filtrado[
        df_mestre_filtrado["inep"] == inep_contato
    ].copy()

    if resultado_inep.empty:
        return None, None, None, None

    candidatos = []

    for _, registro in resultado_inep.iterrows():
        escola_mestre_limpa = registro.get("escola_limpa", "")

        if escola_contato_limpa and escola_mestre_limpa:
            score = fuzz.token_sort_ratio(
                escola_contato_limpa,
                escola_mestre_limpa
            )
        else:
            score = 100

        candidatos.append({
            "registro": registro,
            "score": score
        })

    candidatos = sorted(
        candidatos,
        key=lambda item: item["score"],
        reverse=True
    )

    melhor = candidatos[0]
    registro = melhor["registro"]
    score = melhor["score"]

    if score >= LIMITE_INEP_COM_NOME:
        return (
            int(registro["linha_excel"]),
            registro["escola"],
            "INEP + nome confirmado",
            score
        )

    return None, registro["escola"], "INEP divergente do nome - tentando pelo nome", score


def buscar_por_nome(
    df_mestre_filtrado,
    escola_contato_limpa
):
    if not escola_contato_limpa:
        return None, None, None, None

    escolas_mestre = df_mestre_filtrado["escola_limpa"].tolist()

    if not escolas_mestre:
        return None, None, None, None

    melhor = process.extractOne(
        escola_contato_limpa,
        escolas_mestre,
        scorer=fuzz.token_sort_ratio
    )

    if not melhor:
        return None, None, None, None

    escola_match, score, indice = melhor

    registro_mestre = df_mestre_filtrado.iloc[indice]

    if score >= LIMITE_NOME_ENCONTRADO:
        return (
            int(registro_mestre["linha_excel"]),
            registro_mestre["escola"],
            "Nome da escola",
            score
        )

    if score >= LIMITE_NOME_REVISAR:
        return (
            None,
            registro_mestre["escola"],
            "Possível nome da escola",
            score
        )

    return None, None, None, score


def atualizar_planilha_mestre(
    caminho_planilha_mestre,
    df_contatos,
    cidade_filtro=None
):
    caminho_planilha_mestre = str(caminho_planilha_mestre).strip().strip('"')

    wb = load_workbook(caminho_planilha_mestre)
    ws = localizar_aba_mestre(wb)

    linha_cabecalho = localizar_cabecalho(ws)

    mapa = mapear_colunas_mestre(
        ws,
        linha_cabecalho
    )

    mapa = garantir_colunas_extras(
        ws,
        mapa,
        linha_cabecalho
    )

    df_mestre = carregar_base_mestre(
        ws,
        mapa,
        linha_cabecalho
    )

    if df_mestre.empty:
        raise ValueError(
            "Não foi possível carregar as escolas da planilha mestre."
        )

    if cidade_filtro and cidade_filtro != "Todas":
        cidade_chave = normalizar_chave_cidade(cidade_filtro)

        df_mestre_filtrado = df_mestre[
            df_mestre["cidade_chave"].str.contains(
                cidade_chave,
                na=False
            )
        ].copy()
    else:
        df_mestre_filtrado = df_mestre.copy()

    relatorio = []

    if df_mestre_filtrado.empty:
        relatorio.append({
            "escola_documento": None,
            "inep_documento": None,
            "cidade_documento": cidade_filtro,
            "escola_mestre": None,
            "metodo": None,
            "similaridade": None,
            "gestor_novo": None,
            "telefone_novo": None,
            "telefone_2": None,
            "telefone_3": None,
            "outros_telefones": None,
            "email_novo": None,
            "fonte_contato": None,
            "observacao": None,
            "status": (
                "Nenhuma escola encontrada na planilha mestre para a cidade selecionada. "
                "Verifique o nome da cidade na planilha."
            )
        })

        return pd.DataFrame(relatorio)

    data_atualizacao = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for _, contato in df_contatos.iterrows():
        inep_contato = limpar_inep(contato.get("inep"))
        escola_contato = contato.get("escola")
        cidade_contato = contato.get("municipio")

        gestor_novo = limpar_valor(contato.get("gestor"))
        email_novo = limpar_email(contato.get("email"))

        telefones = coletar_telefones_contato(contato)

        telefone_1 = telefones[0] if len(telefones) >= 1 else ""
        telefone_2 = telefones[1] if len(telefones) >= 2 else ""
        telefone_3 = telefones[2] if len(telefones) >= 3 else ""
        outros_telefones = "; ".join(telefones[3:]) if len(telefones) > 3 else ""

        fonte_contato = limpar_valor(contato.get("fonte_contato"))

        if not fonte_contato:
            fonte_contato = limpar_valor(contato.get("origem"))

        escola_contato_limpa = limpar_nome_escola(escola_contato)

        linha_encontrada = None
        escola_mestre = None
        metodo = None
        similaridade = None
        status = None
        observacao = None

        linha_encontrada, escola_mestre, metodo, similaridade = buscar_por_inep_com_validacao(
            df_mestre_filtrado=df_mestre_filtrado,
            inep_contato=inep_contato,
            escola_contato_limpa=escola_contato_limpa
        )

        if metodo == "INEP divergente do nome - tentando pelo nome":
            observacao = (
                f"O INEP {inep_contato} apareceu no documento, "
                f"mas o nome não bateu com a escola da planilha mestre. "
                f"O sistema tentou localizar pelo nome."
            )
            linha_encontrada = None

        if linha_encontrada is None:
            linha_nome, escola_nome, metodo_nome, similaridade_nome = buscar_por_nome(
                df_mestre_filtrado=df_mestre_filtrado,
                escola_contato_limpa=escola_contato_limpa
            )

            if linha_nome is not None:
                linha_encontrada = linha_nome
                escola_mestre = escola_nome
                metodo = metodo_nome
                similaridade = similaridade_nome

                if observacao:
                    observacao += " Atualizado pelo nome da escola."
            elif metodo_nome == "Possível nome da escola":
                escola_mestre = escola_nome
                metodo = metodo_nome
                similaridade = similaridade_nome
                status = "Revisar manualmente"

        if linha_encontrada is not None:
            if gestor_novo:
                ws.cell(
                    row=linha_encontrada,
                    column=mapa["gestor"]
                ).value = gestor_novo

            if telefone_1:
                ws.cell(
                    row=linha_encontrada,
                    column=mapa["telefone"]
                ).value = telefone_1

            if email_novo:
                ws.cell(
                    row=linha_encontrada,
                    column=mapa["email"]
                ).value = email_novo

            if telefone_2:
                ws.cell(
                    row=linha_encontrada,
                    column=mapa["telefone_2"]
                ).value = telefone_2

            if telefone_3:
                ws.cell(
                    row=linha_encontrada,
                    column=mapa["telefone_3"]
                ).value = telefone_3

            if outros_telefones:
                ws.cell(
                    row=linha_encontrada,
                    column=mapa["outros_telefones"]
                ).value = outros_telefones

            if fonte_contato:
                ws.cell(
                    row=linha_encontrada,
                    column=mapa["fonte_contato"]
                ).value = fonte_contato

            ws.cell(
                row=linha_encontrada,
                column=mapa["data_atualizacao"]
            ).value = data_atualizacao

            status = "Atualizado na planilha mestre"

            ws.cell(
                row=linha_encontrada,
                column=mapa["status_atualizacao"]
            ).value = status

        elif status is None:
            status = "Não encontrado na planilha mestre"

        relatorio.append({
            "escola_documento": escola_contato,
            "inep_documento": inep_contato,
            "cidade_documento": cidade_contato,
            "escola_mestre": escola_mestre,
            "metodo": metodo,
            "similaridade": similaridade,
            "gestor_novo": gestor_novo,
            "telefone_novo": telefone_1,
            "telefone_2": telefone_2,
            "telefone_3": telefone_3,
            "outros_telefones": outros_telefones,
            "email_novo": email_novo,
            "fonte_contato": fonte_contato,
            "observacao": observacao,
            "status": status
        })

    wb.save(caminho_planilha_mestre)

    return pd.DataFrame(relatorio)